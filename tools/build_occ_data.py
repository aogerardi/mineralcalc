#!/usr/bin/env python3
"""Build MineralCalc's Oklahoma well-lookup data from public OCC files.

The OCC publishes daily-refreshed master workbooks (~230 MB combined). This
script downloads them and emits one small JSON per county under data/ok/, so the
site can fetch just the county a user is looking at instead of a giant blob.

Source: https://oklahoma.gov/occ/divisions/oil-gas/oil-gas-data.html
  completions-wells-formations-base.xlsx  — completed wells w/ initial test rates
  ITD-wells-formations-base.xlsx          — intents to drill (permits)

Usage:
    python tools/build_occ_data.py            # download + build
    python tools/build_occ_data.py --no-fetch # rebuild from cached downloads

Output:
    data/ok/index.json      county list + build date, loaded once
    data/ok/<county>.json   wells keyed by section-township-range
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import pathlib
import sys
import urllib.request

BASE = "https://oklahoma.gov/content/dam/ok/en/occ/documents/og/ogdatafiles/"
SOURCES = {
    "completions": "completions-wells-formations-base.xlsx",
    "permits": "ITD-wells-formations-base.xlsx",
}

ROOT = pathlib.Path(__file__).resolve().parent.parent
CACHE = ROOT / "tools" / ".occ-cache"
OUT = ROOT / "data" / "ok"

# Only wells modern enough to inform a horizontal-era estimate. Older wells are
# still on the OCC file but their test rates say nothing useful about a new well.
MIN_YEAR = 2010


def log(msg: str) -> None:
    print(msg, flush=True)


def fetch(name: str, filename: str) -> pathlib.Path:
    CACHE.mkdir(parents=True, exist_ok=True)
    dest = CACHE / filename
    url = BASE + filename
    log(f"  downloading {name}: {url}")
    with urllib.request.urlopen(url, timeout=600) as r, open(dest, "wb") as f:
        while chunk := r.read(1 << 20):
            f.write(chunk)
    log(f"    -> {dest.name} ({dest.stat().st_size / 1e6:.1f} MB)")
    return dest


def norm_county(raw) -> str:
    """'001-ADAIR' and 'ADAIR' both -> 'ADAIR'."""
    s = str(raw or "").strip().upper()
    if "-" in s:
        s = s.split("-", 1)[1]
    return s.strip()


def norm_trs(sec, twp, rng) -> str | None:
    """Canonical section key: '31-06N-05W'."""
    try:
        s = int(str(sec).strip())
    except (TypeError, ValueError):
        return None
    t = str(twp or "").strip().upper().replace(" ", "")
    r = str(rng or "").strip().upper().replace(" ", "")
    if not (1 <= s <= 36) or not t or not r:
        return None
    # Zero-pad the numeric part so keys sort and match consistently (6N -> 06N)
    for part, letters in ((t, "NS"), (r, "EW")):
        if not part[-1] in letters or not part[:-1].isdigit():
            return None
    t = f"{int(t[:-1]):02d}{t[-1]}"
    r = f"{int(r[:-1]):02d}{r[-1]}"
    return f"{s:02d}-{t}-{r}"


def year_of(val) -> int | None:
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.year if 1950 < val.year < 2100 else None
    s = str(val).strip()
    if len(s) >= 4 and s[:4].isdigit():
        y = int(s[:4])
        return y if 1950 < y < 2100 else None
    return None


def ymd(val) -> str | None:
    if isinstance(val, dt.datetime):
        return val.date().isoformat() if 1950 < val.year < 2100 else None
    s = str(val or "").strip()
    if len(s) >= 10 and s[:4].isdigit() and s[4] == "-":
        return s[:10] if s[:4] != "1900" else None
    return None


def num(val) -> float:
    try:
        f = float(val)
        return f if f == f and abs(f) != float("inf") else 0.0
    except (TypeError, ValueError):
        return 0.0


def clean(val) -> str:
    s = str(val or "").strip()
    return "" if s.lower() in ("none", "nan", "") else s


def read_rows(path: pathlib.Path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(it)]
    idx = {name: i for i, name in enumerate(header)}
    for row in it:
        yield row, idx
    wb.close()


def parse_completions(path: pathlib.Path) -> dict:
    """county -> trs -> list of well dicts (deduped by API, best record wins)."""
    counties: dict[str, dict[str, dict]] = {}
    seen: dict[str, tuple] = {}
    total = kept = 0

    for row, ix in read_rows(path):
        total += 1

        def g(col):
            i = ix.get(col)
            return row[i] if i is not None and i < len(row) else None

        # Prefer the bottom-hole location: for a horizontal, that is the section
        # the lateral actually drains, which is what a mineral owner is asking about
        trs = norm_trs(g("BH_Section"), g("BH_Township"), g("BH_Range"))
        county = norm_county(g("BH_County")) if trs else ""
        if not trs:
            trs = norm_trs(g("Section"), g("Township"), g("Range"))
            county = norm_county(g("County"))
        if not trs or not county:
            continue

        first = ymd(g("First_Prod")) or ymd(g("Well_Completion"))
        yr = year_of(first) or year_of(g("Spud"))
        if not yr or yr < MIN_YEAR:
            continue

        api = clean(g("API_Number"))
        oil = round(num(g("Oil_BBL_Per_Day")))
        gas = round(num(g("Gas_MCF_Per_Day")))
        if oil <= 0 and gas <= 0:
            continue  # no test rate = nothing to calibrate against

        well = {
            "api": api,
            "name": (clean(g("Well_Name")) + " " + clean(g("Well_Number"))).strip()[:48],
            "op": clean(g("Operator_Name"))[:44],
            "date": first or f"{yr}-01-01",
            "oil": oil,
            "gas": gas,
            "type": "H" if clean(g("Drill_Type")).upper().startswith("H") else "V",
            "fm": clean(g("Formation_Name"))[:22],
            "lat": round(num(g("Length"))) or None,
            "pun": clean(g("OTC_Prod_Unit_No"))[:16] or None,
        }

        # One well can appear once per formation/casing row — keep the record
        # with the highest reported rate so a multi-row well isn't double counted
        key = api or f"{trs}|{well['name']}"
        rank = (oil * 6 + gas, well["date"])
        prev = seen.get(key)
        if prev and prev[0] >= rank:
            continue
        if prev:
            counties[prev[1]][prev[2]] = [
                w for w in counties[prev[1]][prev[2]] if w.get("api") != api or not api
            ]
        seen[key] = (rank, county, trs)
        counties.setdefault(county, {}).setdefault(trs, []).append(well)
        kept += 1

    log(f"    scanned {total:,} rows, kept {kept:,} completions since {MIN_YEAR}")
    return counties


def parse_permits(path: pathlib.Path) -> dict:
    """county -> trs -> list of recent permit dicts (drilling not yet reported)."""
    counties: dict[str, dict[str, list]] = {}
    cutoff = dt.date.today().year - 3
    total = kept = 0

    for row, ix in read_rows(path):
        total += 1

        def g(col):
            i = ix.get(col)
            return row[i] if i is not None and i < len(row) else None

        if clean(g("Permit_Status")).upper() in ("REJECTED", "CANCELLED", "CANCELED"):
            continue
        trs = norm_trs(g("PBH_Section"), g("PBH_Township"), g("PBH_Range"))
        county = norm_county(g("PBH_County")) if trs else ""
        if not trs:
            trs = norm_trs(g("Section"), g("Township"), g("Range"))
            county = norm_county(g("County"))
        if not trs or not county:
            continue

        appr = ymd(g("Approval_Date")) or ymd(g("Submit_Date"))
        yr = year_of(appr)
        if not yr or yr < cutoff:
            continue

        counties.setdefault(county, {}).setdefault(trs, []).append({
            "api": clean(g("API_Number")),
            "name": (clean(g("Well_Name")) + " " + clean(g("Well_Number"))).strip()[:48],
            "op": clean(g("Entity_Name"))[:44],
            "date": appr,
            "fm": clean(g("Formation_Name"))[:22],
            "type": "H" if clean(g("Drill_Type")).upper().startswith("H") else "V",
        })
        kept += 1

    log(f"    scanned {total:,} rows, kept {kept:,} permits since {cutoff}")
    return counties


def slug(county: str) -> str:
    return county.lower().replace(" ", "-")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-fetch", action="store_true", help="use cached downloads")
    args = ap.parse_args()

    paths = {}
    for name, filename in SOURCES.items():
        cached = CACHE / filename
        if args.no_fetch:
            if not cached.exists():
                log(f"ERROR: --no-fetch but {cached} is missing")
                return 1
            paths[name] = cached
        else:
            paths[name] = fetch(name, filename)

    log("  parsing completions...")
    comps = parse_completions(paths["completions"])
    log("  parsing permits...")
    perms = parse_permits(paths["permits"])

    OUT.mkdir(parents=True, exist_ok=True)
    for stale in OUT.glob("*.json"):
        stale.unlink()

    today = dt.date.today().isoformat()
    counties = sorted(set(comps) | set(perms))
    index = {"updated": today, "minYear": MIN_YEAR, "counties": []}
    total_bytes = 0

    for county in counties:
        if not county or county == "NONE":
            continue
        sections = {}
        for trs, wells in comps.get(county, {}).items():
            wells.sort(key=lambda w: w["date"], reverse=True)
            sections.setdefault(trs, {})["w"] = wells[:40]
        for trs, permits in perms.get(county, {}).items():
            permits.sort(key=lambda p: p["date"], reverse=True)
            sections.setdefault(trs, {})["p"] = permits[:20]
        if not sections:
            continue
        payload = {"county": county, "updated": today, "sections": sections}
        path = OUT / f"{slug(county)}.json"
        path.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
        size = path.stat().st_size
        total_bytes += size
        index["counties"].append({
            "name": county,
            "file": path.name,
            "sections": len(sections),
            "kb": round(size / 1024),
        })

    (OUT / "index.json").write_text(json.dumps(index, separators=(",", ":")), encoding="utf-8")
    log(f"\nBuilt {len(index['counties'])} counties, {total_bytes / 1e6:.1f} MB total")
    log(f"Largest: " + ", ".join(
        f"{c['name']} {c['kb']}KB"
        for c in sorted(index["counties"], key=lambda c: -c["kb"])[:5]))
    return 0


if __name__ == "__main__":
    sys.exit(main())
